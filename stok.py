#Dvice was here

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import csv
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.patches as mpatches

class StockApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Dvice Stok Takip Uygulaması")
        self.root.geometry("1200x750")
        self.root.configure(bg="#f8f9fa")

        self.db_connection = sqlite3.connect("stock.db")
        self.create_table()

        self.products = {}
        self.load_data_from_db()
        self.create_ui()

    def create_table(self):
        query = """
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY,
            name TEXT,
            quantity INTEGER,
            min_stock INTEGER
        );
        """
        self.db_connection.execute(query)
        self.db_connection.commit()

    def load_data_from_db(self):
        query = "SELECT * FROM products"
        cursor = self.db_connection.execute(query)
        for row in cursor:
            self.products[row[1]] = {'quantity': row[2], 'min_stock': row[3]}

    def save_data_to_db(self):
        query = "DELETE FROM products"
        self.db_connection.execute(query)
        for product, data in self.products.items():
            query = "INSERT INTO products (name, quantity, min_stock) VALUES (?, ?, ?)"
            self.db_connection.execute(query, (product, data['quantity'], data['min_stock']))
        self.db_connection.commit()

    def create_ui(self):
        # Ana çerçeve
        main_frame = tk.Frame(self.root, bg="#f8f9fa")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Sol panel - Kontroller
        left_frame = tk.LabelFrame(main_frame, text="Stok İşlemleri", bg="#ffffff", fg="#2c3e50", 
                                  font=("Segoe UI", 10, "bold"), padx=15, pady=15)
        left_frame.pack(side="left", fill="y", padx=(0, 20))

        # Ürün Ekle Bölümü
        add_frame = tk.LabelFrame(left_frame, text="Ürün Ekle", bg="#ffffff", fg="#34495e", 
                                 font=("Segoe UI", 9, "bold"))
        add_frame.pack(fill="x", pady=(0, 15))

        tk.Label(add_frame, text="Ürün Adı:", bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.entry_name = tk.Entry(add_frame, font=("Segoe UI", 9), relief="solid", bd=1)
        self.entry_name.grid(row=0, column=1, sticky="ew", padx=5, pady=5)

        tk.Label(add_frame, text="Stok Miktarı:", bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.entry_quantity = tk.Entry(add_frame, font=("Segoe UI", 9), relief="solid", bd=1)
        self.entry_quantity.grid(row=1, column=1, sticky="ew", padx=5, pady=5)

        tk.Label(add_frame, text="Minimum Stok:", bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 9)).grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.entry_min_stock = tk.Entry(add_frame, font=("Segoe UI", 9), relief="solid", bd=1)
        self.entry_min_stock.grid(row=2, column=1, sticky="ew", padx=5, pady=5)

        add_frame.columnconfigure(1, weight=1)

        self.add_button = tk.Button(add_frame, text="Ürün Ekle", command=self.add_product, 
                                   bg="#27ae60", fg="white", font=("Segoe UI", 9, "bold"),
                                   relief="flat", cursor="hand2", activebackground="#2ecc71")
        self.add_button.grid(row=3, column=0, columnspan=2, sticky="ew", padx=5, pady=(10, 5))

        # Stok İşlemleri Bölümü
        operations_frame = tk.LabelFrame(left_frame, text="Stok İşlemleri", bg="#ffffff", fg="#34495e", 
                                        font=("Segoe UI", 9, "bold"))
        operations_frame.pack(fill="x", pady=(0, 15))

        self.show_button = tk.Button(operations_frame, text="Detaylı Stok Listesi", command=self.show_stock_list_window, 
                                    bg="#3498db", fg="white", font=("Segoe UI", 9),
                                    relief="flat", cursor="hand2", activebackground="#5dade2")
        self.show_button.pack(fill="x", padx=5, pady=5)

        self.show_total_stock_button = tk.Button(operations_frame, text="Toplam Stok Görüntüle", command=self.show_total_stock, 
                                                bg="#8e44ad", fg="white", font=("Segoe UI", 9),
                                                relief="flat", cursor="hand2", activebackground="#a569bd")
        self.show_total_stock_button.pack(fill="x", padx=5, pady=5)

        # Arama ve Filtreleme
        filter_frame = tk.LabelFrame(left_frame, text="Arama ve Filtreleme", bg="#ffffff", fg="#34495e", 
                                    font=("Segoe UI", 9, "bold"))
        filter_frame.pack(fill="x", pady=(0, 15))

        tk.Label(filter_frame, text="Ürün Ara:", bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=(5,0))
        self.entry_filter = tk.Entry(filter_frame, font=("Segoe UI", 9), relief="solid", bd=1)
        self.entry_filter.pack(fill="x", padx=5, pady=5)
        
        self.filter_button = tk.Button(filter_frame, text="Filtrele", command=self.filter_products, 
                                      bg="#f39c12", fg="white", font=("Segoe UI", 9),
                                      relief="flat", cursor="hand2", activebackground="#f4d03f")
        self.filter_button.pack(fill="x", padx=5, pady=5)

        # Stok Azaltma Bölümü
        decrease_frame = tk.LabelFrame(left_frame, text="Stok Azalt", bg="#ffffff", fg="#34495e", 
                                      font=("Segoe UI", 9, "bold"))
        decrease_frame.pack(fill="x", pady=(0, 15))

        tk.Label(decrease_frame, text="Ürün Adı:", bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.entry_decrease = tk.Entry(decrease_frame, font=("Segoe UI", 9), relief="solid", bd=1)
        self.entry_decrease.grid(row=0, column=1, sticky="ew", padx=5, pady=5)

        tk.Label(decrease_frame, text="Miktar:", bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.entry_decrease_amount = tk.Entry(decrease_frame, font=("Segoe UI", 9), relief="solid", bd=1)
        self.entry_decrease_amount.grid(row=1, column=1, sticky="ew", padx=5, pady=5)

        decrease_frame.columnconfigure(1, weight=1)

        self.decrease_button = tk.Button(decrease_frame, text="Stok Azalt", command=self.decrease_stock, 
                                        bg="#e74c3c", fg="white", font=("Segoe UI", 9),
                                        relief="flat", cursor="hand2", activebackground="#ec7063")
        self.decrease_button.grid(row=2, column=0, columnspan=2, sticky="ew", padx=5, pady=(10, 5))

        # Ürün Silme Bölümü
        delete_frame = tk.LabelFrame(left_frame, text="Ürün Sil", bg="#ffffff", fg="#34495e", 
                                    font=("Segoe UI", 9, "bold"))
        delete_frame.pack(fill="x")

        tk.Label(delete_frame, text="Ürün Adı:", bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=(5,0))
        self.entry_delete = tk.Entry(delete_frame, font=("Segoe UI", 9), relief="solid", bd=1)
        self.entry_delete.pack(fill="x", padx=5, pady=5)

        self.delete_button = tk.Button(delete_frame, text="Ürünü Sil", command=self.delete_product, 
                                      bg="#c0392b", fg="white", font=("Segoe UI", 9),
                                      relief="flat", cursor="hand2", activebackground="#d5423e")
        self.delete_button.pack(fill="x", padx=5, pady=5)

        #  Tablo ve Grafik
        right_frame = tk.Frame(main_frame, bg="#f8f9fa")
        right_frame.pack(side="right", fill="both", expand=True)

        #  Stok Durumu Özeti
        summary_frame = tk.LabelFrame(right_frame, text="Stok Durumu Özeti", bg="#ffffff", fg="#2c3e50", 
                                     font=("Segoe UI", 10, "bold"), padx=10, pady=10)
        summary_frame.pack(fill="x", pady=(0, 10))

        self.total_stock_label = tk.Label(summary_frame, text="Toplam Stok: 0", bg="#ffffff", fg="#2c3e50", 
                                         font=("Segoe UI", 11, "bold"))
        self.total_stock_label.pack(side="left")

        self.critical_stock_label = tk.Label(summary_frame, text="Kritik Stok: 0", bg="#ffffff", fg="#e74c3c", 
                                           font=("Segoe UI", 11, "bold"))
        self.critical_stock_label.pack(side="right")

        # Orta kısım - Tablo
        table_frame = tk.LabelFrame(right_frame, text="Stok Listesi", bg="#ffffff", fg="#2c3e50", 
                                   font=("Segoe UI", 10, "bold"))
        table_frame.pack(fill="both", expand=True, pady=(0, 10))

        self.tree = ttk.Treeview(table_frame, columns=("Ürün", "Stok Miktarı", "Min. Stok", "Durum"), show="headings")
        self.tree.heading("Ürün", text="Ürün")
        self.tree.heading("Stok Miktarı", text="Stok Miktarı")
        self.tree.heading("Min. Stok", text="Min. Stok")
        self.tree.heading("Durum", text="Durum %")
        
        for col in ("Ürün", "Stok Miktarı", "Min. Stok", "Durum"):
            self.tree.column(col, anchor="center", width=120)
        
        scrollbar_table = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar_table.set)
        
        self.tree.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scrollbar_table.pack(side="right", fill="y")

        # Grafik
        chart_frame = tk.LabelFrame(right_frame, text="Stok Durumu Grafiği", bg="#ffffff", fg="#2c3e50", 
                                   font=("Segoe UI", 10, "bold"))
        chart_frame.pack(fill="x", pady=(0, 0))


        self.fig, self.ax = plt.subplots(figsize=(8, 3), facecolor='white')
        self.canvas = FigureCanvasTkAgg(self.fig, chart_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=5, pady=5)

        self.update_table()
        self.update_chart()

    def add_product(self):
        product_name = self.entry_name.get().strip()
        try:
            quantity = int(self.entry_quantity.get())
            min_stock = int(self.entry_min_stock.get())
        except ValueError:
            messagebox.showerror("Hata", "Miktar ve minimum stok sayı olmalı!")
            return

        if product_name and quantity >= 0 and min_stock >= 0:
            if product_name in self.products:
                self.products[product_name]['quantity'] += quantity
            else:
                self.products[product_name] = {'quantity': quantity, 'min_stock': min_stock}

            messagebox.showinfo("Başarılı", f"{product_name} stoklara eklendi.")
            self.entry_name.delete(0, tk.END)
            self.entry_quantity.delete(0, tk.END)
            self.entry_min_stock.delete(0, tk.END)
            self.update_table()
            self.update_chart()
        else:
            messagebox.showerror("Hata", "Geçersiz giriş!")

    def show_stock_list_window(self):
        window = tk.Toplevel(self.root)
        window.title("Detaylı Stok Listesi")
        window.geometry("1000x600")
        window.configure(bg="#f8f9fa")

        main_frame = tk.Frame(window, bg="#f8f9fa")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Sol taraf - Tablo
        left_panel = tk.LabelFrame(main_frame, text="Ürün Listesi", bg="#ffffff", fg="#2c3e50", 
                                  font=("Segoe UI", 10, "bold"))
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 10))

        columns = ("Ürün", "Stok Miktarı", "Min. Stok", "Durum")
        tree = ttk.Treeview(left_panel, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=150)
        tree.pack(side="left", expand=True, fill="both", padx=10, pady=10)

        scrollbar = ttk.Scrollbar(left_panel, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        for product, data in self.products.items():
            percentage = (data['quantity'] / max(data['min_stock'], 1)) * 100
            status = f"%{percentage:.1f}"
            item = tree.insert("", "end", values=(product, data['quantity'], data['min_stock'], status))
            
            if data['quantity'] < data['min_stock']:
                tree.item(item, tags=("critical",))
            elif data['quantity'] < data['min_stock'] * 1.5:
                tree.item(item, tags=("warning",))
            else:
                tree.item(item, tags=("good",))

        tree.tag_configure("critical", background="#e74c3c", foreground="white")
        tree.tag_configure("warning", background="#f39c12", foreground="white")
        tree.tag_configure("good", background="#27ae60", foreground="white")

        #Grafik ve Kontroller
        right_panel = tk.Frame(main_frame, bg="#f8f9fa")
        right_panel.pack(side="right", fill="both", padx=(10, 0))

        # Grafik
        chart_frame = tk.LabelFrame(right_panel, text="Stok Yüzde Grafiği", bg="#ffffff", fg="#2c3e50", 
                                   font=("Segoe UI", 10, "bold"))
        chart_frame.pack(fill="both", expand=True, pady=(0, 10))

        fig, ax = plt.subplots(figsize=(6, 4), facecolor='white')
        canvas = FigureCanvasTkAgg(fig, chart_frame)
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=5, pady=5)

        if self.products:
            products = list(self.products.keys())[:10]  # İlk 10 ürün
            percentages = []
            colors = []
            
            for product in products:
                data = self.products[product]
                percentage = (data['quantity'] / max(data['min_stock'], 1)) * 100
                percentages.append(percentage)
                
                if percentage < 100:
                    colors.append('#e74c3c')  
                elif percentage < 150:
                    colors.append('#f39c12')  
                else:
                    colors.append('#27ae60') 

            bars = ax.bar(range(len(products)), percentages, color=colors, alpha=0.7)
            ax.set_xlabel('Ürünler', fontsize=10)
            ax.set_ylabel('Stok Yüzdesi (%)', fontsize=10)
            ax.set_title('Ürün Stok Yüzdeleri', fontsize=12, fontweight='bold')
            ax.set_xticks(range(len(products)))
            ax.set_xticklabels([p[:8] + '...' if len(p) > 8 else p for p in products], rotation=45)
            ax.axhline(y=100, color='red', linestyle='--', alpha=0.5, label='Minimum Stok Seviyesi')
            ax.legend()
            ax.grid(True, alpha=0.3)

            # Değerleri barların üstüne yaz
            for i, (bar, percentage) in enumerate(zip(bars, percentages)):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height + 5,
                       f'{percentage:.0f}%', ha='center', va='bottom', fontsize=8)

        plt.tight_layout()
        canvas.draw()

        # Kontrol paneli
        controls_frame = tk.LabelFrame(right_panel, text="İşlemler", bg="#ffffff", fg="#2c3e50", 
                                      font=("Segoe UI", 10, "bold"))
        controls_frame.pack(fill="x")

        # CSV 
        def export_csv():
            # Dosya formatı seçimi
            file_types = [("Excel Dosyası", "*.xlsx"), ("CSV Dosyası (Excel Uyumlu)", "*.csv"), ("Metin Dosyası", "*.txt")]
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=file_types)
            if path:
                try:
                    if path.endswith('.xlsx'):
                        # Excel dosyası olarak kaydet 
                        import openpyxl
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Stok Listesi"
                        
                        # Başlıklar
                        headers = ["Ürün", "Stok Miktarı", "Min. Stok", "Durum %", "Stok Durumu"]
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=header)
                        
                        # Veriler
                        for row, (product, data) in enumerate(self.products.items(), 2):
                            percentage = (data['quantity'] / max(data['min_stock'], 1)) * 100
                            status = "Normal" if percentage >= 100 else "Kritik"
                            ws.cell(row=row, column=1, value=product)
                            ws.cell(row=row, column=2, value=data['quantity'])
                            ws.cell(row=row, column=3, value=data['min_stock'])
                            ws.cell(row=row, column=4, value=f"{percentage:.1f}%")
                            ws.cell(row=row, column=5, value=status)
                        
                        wb.save(path)
                        messagebox.showinfo("Başarılı", f"Excel dosyası kaydedildi:\n{path}\n")
                        
                    else:
                        # CSV için Excel uyumlu format
                        encoding = "utf-8-sig" 
                        with open(path, mode="w", newline="", encoding=encoding) as f:
                            writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_ALL)  
                            writer.writerow(["Ürün", "Stok Miktarı", "Min. Stok", "Durum %", "Stok Durumu"])
                            for product, data in self.products.items():
                                percentage = (data['quantity'] / max(data['min_stock'], 1)) * 100
                                status = "Normal" if percentage >= 100 else "Kritik"
                                writer.writerow([product, data['quantity'], data['min_stock'], f"{percentage:.1f}%", status])
                        
                        messagebox.showinfo("Başarılı", f"CSV kaydedildi:\n{path}\n\nExcel'de açarken 'Veri > Metinden Sütunlara' kullanın\nveya dosya uzantısını .xlsx yapın.")
                        
                except ImportError:
                    # openpyxl yoksa CSV'ye geçme
                    messagebox.showwarning("Uyarı", "Excel desteği için 'pip install openpyxl' çalıştırın.\nCSV olarak kaydediliyor...")
                    encoding = "utf-8-sig"
                    with open(path.replace('.xlsx', '.csv'), mode="w", newline="", encoding=encoding) as f:
                        writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_ALL)
                        writer.writerow(["Ürün", "Stok Miktarı", "Min. Stok", "Durum %"])
                        for product, data in self.products.items():
                            percentage = (data['quantity'] / max(data['min_stock'], 1)) * 100
                            writer.writerow([product, data['quantity'], data['min_stock'], f"{percentage:.1f}%"])
                except Exception as e:
                    messagebox.showerror("Hata", f"Dosya kaydedilemedi!\n{e}")

        export_button = tk.Button(controls_frame, text="CSV'ye Aktar", command=export_csv, 
                                 bg="#16a085", fg="white", font=("Segoe UI", 9),
                                 relief="flat", cursor="hand2", activebackground="#48c9b0")
        export_button.pack(fill="x", padx=5, pady=5)

        # Sil butonu
        delete_button = tk.Button(controls_frame, text="Seçili Ürünü Sil", state="disabled", 
                                 bg="#c0392b", fg="white", font=("Segoe UI", 9),
                                 relief="flat", cursor="hand2", activebackground="#d5423e")
        delete_button.pack(fill="x", padx=5, pady=5)

        # Seçim olayı
        def on_tree_select(event):
            selected = tree.selection()
            delete_button.config(state="normal" if selected else "disabled")

        tree.bind("<<TreeviewSelect>>", on_tree_select)

        # Silme işlemi
        def delete_selected():
            selected = tree.selection()
            if not selected:
                return
            item = selected[0]
            product_name = tree.item(item, "values")[0]
            if messagebox.askyesno("Onay", f"{product_name} ürününü silmek istediğinize emin misiniz?"):
                if product_name in self.products:
                    del self.products[product_name]
                    self.update_table()
                    self.update_chart()
                tree.delete(item)

        delete_button.config(command=delete_selected)

    def filter_products(self):
        filter_text = self.entry_filter.get().strip()
        if not filter_text:
            messagebox.showwarning("Uyarı", "Lütfen bir arama terimi girin!")
            return
            
        filtered_text = "Filtrelenmiş Ürünler:\n\n"
        found_count = 0
        
        for product, data in self.products.items():
            if filter_text.lower() in product.lower():
                percentage = (data['quantity'] / max(data['min_stock'], 1)) * 100
                status = "✅ Normal" if percentage >= 100 else "⚠️ Kritik"
                filtered_text += f"• {product}: {data['quantity']} adet (Min: {data['min_stock']}) - {status}\n"
                found_count += 1
        
        if found_count == 0:
            filtered_text += "Eşleşen ürün bulunamadı."
        else:
            filtered_text = f"Bulunan ürün sayısı: {found_count}\n\n" + filtered_text
            
        messagebox.showinfo("Arama Sonuçları", filtered_text)

    def show_total_stock(self):
        total_stock = sum(data['quantity'] for data in self.products.values())
        critical_count = sum(1 for data in self.products.values() if data['quantity'] < data['min_stock'])
        
        info_text = f"Toplam Stok Miktarı: {total_stock}\n"
        info_text += f"Toplam Ürün Çeşidi: {len(self.products)}\n"
        info_text += f"Kritik Seviyedeki Ürünler: {critical_count}"
        
        self.total_stock_label.config(text=f"Toplam Stok: {total_stock}")
        self.critical_stock_label.config(text=f"Kritik Stok: {critical_count}")
        messagebox.showinfo("Stok İstatistikleri", info_text)

    def delete_product(self):
        product_name = self.entry_delete.get().strip()
        if product_name in self.products:
            if messagebox.askyesno("Onay", f"{product_name} ürününü silmek istediğinize emin misiniz?"):
                del self.products[product_name]
                messagebox.showinfo("Başarılı", f"{product_name} stoklardan silindi.")
                self.entry_delete.delete(0, tk.END)
                self.update_table()
                self.update_chart()
        else:
            messagebox.showerror("Hata", f"{product_name} bulunamadı!")

    def decrease_stock(self):
        product_name = self.entry_decrease.get().strip()
        try:
            decrease_amount = int(self.entry_decrease_amount.get())
        except ValueError:
            messagebox.showerror("Hata", "Eksilme miktarı sayı olmalı!")
            return

        if product_name in self.products:
            if self.products[product_name]['quantity'] >= decrease_amount:
                self.products[product_name]['quantity'] -= decrease_amount
                messagebox.showinfo("Başarılı", f"{product_name} stok azaltıldı.")
                self.entry_decrease.delete(0, tk.END)
                self.entry_decrease_amount.delete(0, tk.END)
                self.update_table()
                self.update_chart()
                
                # Kritik kontrolü
                if self.products[product_name]['quantity'] < self.products[product_name]['min_stock']:
                    messagebox.showwarning("Uyarı", f"{product_name} minimum stok seviyesinin altına düştü!")
            else:
                messagebox.showerror("Hata", f"{product_name} stokta yeterli miktar yok!")
        else:
            messagebox.showerror("Hata", f"{product_name} bulunamadı!")

    def update_table(self):
        self.tree.delete(*self.tree.get_children())
        for product, data in self.products.items():
            percentage = (data['quantity'] / max(data['min_stock'], 1)) * 100
            status_text = f"%{percentage:.1f}"
            
            item = self.tree.insert("", "end", values=(product, data['quantity'], data['min_stock'], status_text))
            
            if data['quantity'] < data['min_stock']:
                self.tree.item(item, tags=("critical",))
            elif data['quantity'] < data['min_stock'] * 1.5:
                self.tree.item(item, tags=("warning",))
            else:
                self.tree.item(item, tags=("good",))

        self.tree.tag_configure("critical", background="#e74c3c", foreground="white")
        self.tree.tag_configure("warning", background="#f39c12", foreground="white")
        self.tree.tag_configure("good", background="#27ae60", foreground="white")

        # Özet bilgileri güncelle
        total_stock = sum(data['quantity'] for data in self.products.values())
        critical_count = sum(1 for data in self.products.values() if data['quantity'] < data['min_stock'])
        self.total_stock_label.config(text=f"Toplam Stok: {total_stock}")
        self.critical_stock_label.config(text=f"Kritik Stok: {critical_count}")

    def update_chart(self):
        self.ax.clear()
        
        if not self.products:
            self.ax.text(0.5, 0.5, 'Henüz ürün eklenmemiş', ha='center', va='center', 
                        transform=self.ax.transAxes, fontsize=12, color='gray')
            self.canvas.draw()
            return

        # En fazla 10 ürün göster
        products = list(self.products.items())[:10]
        product_names = [p[0] for p in products]
        percentages = []
        colors = []
        
        for product_name, data in products:
            percentage = (data['quantity'] / max(data['min_stock'], 1)) * 100
            percentages.append(percentage)
            
            if percentage < 100:
                colors.append('#e74c3c')  # Kırmızı - Kritik Stok
            elif percentage < 150:
                colors.append('#f39c12')  # Turuncu - Uyarı almış Stok
            else:
                colors.append('#27ae60')  # Yeşil - Normal Stok


        bars = self.ax.bar(range(len(product_names)), percentages, color=colors, alpha=0.8, edgecolor='white', linewidth=1)
        
        self.ax.set_xlabel('Ürünler', fontsize=10, color='#2c3e50')
        self.ax.set_ylabel('Minimum Stoka Göre Yüzde (%)', fontsize=10, color='#2c3e50')
        self.ax.set_title('Stok Durumu - Yüzdelik Gösterim', fontsize=12, fontweight='bold', color='#2c3e50', pad=20)
        

        self.ax.set_xticks(range(len(product_names)))
        self.ax.set_xticklabels([name[:10] + '...' if len(name) > 10 else name for name in product_names], 
                               rotation=45, ha='right', fontsize=9)
        
        # Yüzde çizgisi
        self.ax.axhline(y=100, color='#e74c3c', linestyle='--', alpha=0.7, linewidth=2, label='Min. Stok Seviyesi')
        self.ax.axhline(y=150, color='#f39c12', linestyle='--', alpha=0.5, linewidth=1, label='İdeal Stok Seviyesi')
        

        critical_patch = mpatches.Patch(color='#e74c3c', label='Kritik (<100%)')
        warning_patch = mpatches.Patch(color='#f39c12', label='Uyarı (100-150%)')
        good_patch = mpatches.Patch(color='#27ae60', label='Normal (>150%)')
        self.ax.legend(handles=[critical_patch, warning_patch, good_patch], loc='upper right', fontsize=8)
        
        # Değerleri barların üstüne yaz
        for i, (bar, percentage) in enumerate(zip(bars, percentages)):
            height = bar.get_height()
            self.ax.text(bar.get_x() + bar.get_width()/2., height + max(percentages)*0.02,
                        f'{percentage:.0f}%', ha='center', va='bottom', fontsize=8, fontweight='bold')
        
        self.ax.grid(True, alpha=0.2, axis='y')
        self.ax.set_facecolor('#fafafa')
        
        plt.tight_layout()
        self.canvas.draw()

    def __del__(self):
        self.save_data_to_db()

if __name__ == "__main__":
    root = tk.Tk()
    app = StockApp(root)
    root.mainloop()


#Dvice was here